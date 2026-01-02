"""Excel Parser using openpyxl"""
from openpyxl import load_workbook
from typing import Dict, List
from loguru import logger


class ExcelParser:
    """Parse Excel files and extract text"""
    
    def parse(self, file_path: str) -> Dict:
        """Parse Excel and extract content from all sheets"""
        try:
            workbook = load_workbook(file_path, data_only=True)
            logger.info(f"Opened Excel: {file_path}, {len(workbook.sheetnames)} sheets")
            
            sheets_data = []
            all_text = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = []
                
                # Extract all cell values
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        sheet_text.append(' | '.join(row_text))
                
                sheet_content = '\n'.join(sheet_text)
                sheets_data.append({
                    'name': sheet_name,
                    'content': sheet_content,
                })
                all_text.append(f"Sheet: {sheet_name}\n{sheet_content}")
            
            workbook.close()
            
            full_content = '\n\n---\n\n'.join(all_text)
            
            return {
                'total_pages': len(workbook.sheetnames),
                'chapters': [{
                    'number': None,
                    'title': 'Excel Content',
                    'start_page': 1,
                    'end_page': len(workbook.sheetnames),
                    'content': full_content,
                }],
                'metadata': {
                    'title': workbook.properties.title or '',
                    'author': workbook.properties.creator or '',
                },
                'file_type': 'excel',
                'sheets': sheets_data,
            }
            
        except Exception as e:
            logger.error(f"Error parsing Excel {file_path}: {e}")
            raise

